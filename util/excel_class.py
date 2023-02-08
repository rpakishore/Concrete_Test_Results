import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np

pd.options.mode.chained_assignment = None  # default='warn'
class xlsx():
    def __init__(self, file):
        self.wb = openpyxl.Workbook()
        self.wb.active.title = "Summary"
        self.file = file
        self.saved = False
        self.define_styles()

    def define_styles(self):
        self.fonts = {
            'h1': Font(name='Arial', size=12, italic=False, bold=True, color='FF000000', strike=False,underline='none',vertAlign=None)
        }

        self.alignments = {
            'center':Alignment(horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
        }

        self.borderthickness = {
            'thin': Side(border_style="thin", color="000000"),
            'double': Side(border_style="double", color="000000"),
            'thick': Side(border_style="thick", color="000000")
        }
    def write_excel_data(self, ws, data, start_row, start_col = 1):
        """Writes a array of array to excel worksheet

        Args:
            ws (Openpyxl Worksheet Class): Sheet to which the data is written
            data (Array of Arrays): Data to write in array of array format; [[R1C1, R1C2],[R2C1, R2C2]]
            start_row (int): Start row; 1-indexed;Excel row to write to
        """
        #print(f"Writing {len(data)} rows of data to {ws}, starting at {start_row}")
        for i in range(start_row, start_row + len(data)):
            for j in range(start_col + len(data[i-start_row]) - 1):                
                cell = (openpyxl.utils.get_column_letter(1+j)) + str(i)
                ws[cell] = data[i-start_row][j]
                self.saved = False
        
        return start_row + len(data)

    def draw_border(self, ws, data, start_row, start_col=1):
        max_rows = len(data) + start_row
        max_cols = start_col

        for row in data:
            max_cols = max(max_cols, len(row)+start_col)

        for i in range(start_row, max_rows):
            for j in range(start_col-1, max_cols-1):
                cell = openpyxl.utils.get_column_letter(j+1) + str(i)
                left = self.borderthickness['thin']
                right = self.borderthickness['thin']
                top = self.borderthickness['thin']
                bottom = self.borderthickness['thin']
                if i == start_row:
                    top = self.borderthickness['thick']
                if i == max_rows-1:
                    bottom = self.borderthickness['thick']
                if j == start_col-1:
                    left = self.borderthickness['thick']
                if j == max_cols-2:
                    right = self.borderthickness['thick']
                
                ws[cell].border = Border(left=left, right=right, top=top, bottom=bottom)
                ws[cell].alignment = Alignment(wrap_text=True, vertical="center")

    def create_data_sheet(self, title, sample_data):
        ws = self.wb.create_sheet(index=0, title=title)
        self.saved = False
        data = [
            ['Powell River WWTP'],
            ['Concrete Compressive Strength Summary'],
            [f"{sample_data['Mix Number']}: {sample_data['Specified Strength']}MPa @ {sample_data['Specified Strength Days']} days"],
            ['']
        ]
        next_row = self.write_excel_data(ws, data, 1)
        data = [
            ['Set', 'Cylinder', 'Description', 'Date', 'Date', 'Date', 'Age', 'Test', r'% of', 'Air', 'Slump', 'Comments'],
            ['#', 'ID', '', 'Cast', 'Received', 'Tested', 'Days', 'MPa', 'Design', '%', 'mm', '']
        ]
        self.draw_border(ws, data,next_row)
        next_row = self.write_excel_data(ws, data, next_row)

        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["L"].width = 30
        
        for cells in ws["1:3"]:
            for cell in cells:
                cell.font = self.fonts['h1']
        for cells in ws["4:6"]:
            for cell in cells:
                cell.alignment = self.alignments['center']

        ws.freeze_panes = 'A7'
        return ws, next_row

    def write_set_data(self, ws, data, startrow,startcol=1):
        data_to_write = []

        for index, row in data['Test Data'].iterrows():
            if index == 0:
                location_comment = [x.strip().title() for x in data['Location Comments']]
                other_comment = [x.strip().title() for x in data['Other Comments']]
                data_to_write.append([
                    data['Set Num'], 
                    row['Specimen'], 
                    '\n'.join(location_comment),
                    datetime.strftime(data['Cast Date'], "%d-%B-%Y"),
                    datetime.strftime(data['Transported Date'], "%d-%B-%Y"),
                    datetime.strftime(row['Test_Date'], "%d-%B-%Y"),
                    int(row['Age']) if row['Age'] != None else None,
                    float(row['Compressive_Str']) if row['Compressive_Str'] != None else None,
                    round(float(row['Compressive_Str']) / data['Specified Strength'],2) if row['Compressive_Str'] != None else None,
                    data['Air'],
                    data['Slump'],
                    '\n'.join(other_comment)
                    ])
            else:
                data_to_write.append([
                    '', '', '', '', '', '',
                    int(row['Age']) if row['Age'] != None else None,
                    float(row['Compressive_Str']) if row['Compressive_Str'] != None else None,
                    round(float(row['Compressive_Str']) / data['Specified Strength'],2) if row['Compressive_Str'] != None else None,
                ])


        next_row = self.write_excel_data(ws, data_to_write, startrow)
        self.draw_border(ws, data_to_write, startrow)
        return next_row

    def plot_sheet(self, sets, filename):
        x_axis = []
        req =[]
        str_days = []
        for set_key in sorted(sets.keys()):
            for index, row in sets[set_key].extracted_data['Test Data'].iterrows():
                if row['Compressive_Str'] and int(row['Age']) <= sets[set_key].extracted_data['Specified Strength Days']: 
                        if not int(row['Age']) in str_days:
                            str_days.append(int(row['Age']))

        df = pd.DataFrame(columns=sorted(str_days))

        for set_key in sorted(sets.keys()):
            #Create dataframe with setwise data      
            df.loc[set_key] = 0  
            for index, row in sets[set_key].extracted_data['Test Data'].iterrows():
                if row['Compressive_Str'] and int(row['Age']) <= sets[set_key].extracted_data['Specified Strength Days']:
                        df[int(row['Age'])][set_key] = float(row['Compressive_Str'])

            str_days = sorted(str_days)
            x_axis.append(str(set_key))
            req.append(sets[set_key].extracted_data['Specified Strength'])

        plt.clf()
        plt.bar(x_axis, req, color='r', label='Target Str.')
        plt.axhline(y=req[0],linewidth=1, color='r', linestyle='--')
        c = np.arange(1, len(str_days) + 1)
        norm = mpl.colors.Normalize(vmin=c.min(), vmax=c.max())
        cmap = mpl.cm.ScalarMappable(norm=norm, cmap=mpl.cm.Blues)
        cmap.set_array([])
        i=0
        for index, row in df.transpose().iterrows():
            plt.bar(x_axis, row.to_list(), width=0.8, color=cmap.to_rgba(i + 1), alpha=1, label=f"{row.name} days")
            i+=1
        plt.legend(bbox_to_anchor=(1.04,1), loc="upper left")
        plt.xlabel('Set #')
        plt.ylabel('MPa')
        plt.title("How to read: You shouldn't see any red bars in graph", fontsize=8)
        
        target={'str':sets[set_key].extracted_data['Specified Strength'], 'day': sets[set_key].extracted_data['Specified Strength Days']}
        plt.suptitle(f"{sets[set_key].extracted_data['Mix Number']}: {target['str']}MPa @ {target['day']} days test summary")
        plt.savefig(filename,bbox_inches='tight')
        return filename
        #plt.plot()

    def save(self):
        self.wb.save(self.file)
        self.saved = True
